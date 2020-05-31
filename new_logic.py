import openpyxl as xl
import itertools

wb = xl.load_workbook("dataa2.xlsx")
sheet1 = wb['flow handoff']
sheet2 = wb['mpstat']
count1 = 0
count2 = 0
date = []
max_r1 = sheet2.max_row
for r1 in range(2, max_r1 + 1):
    date_ele = sheet2.cell(r1, 1)
    date.append(date_ele.value)

# print (date)
i = 0
max_r2 = sheet1.max_row
for row in sheet2.iter_rows():
    for cell in row:
        if cell.value != date[i]:
            sheet1.insert_rows(row, amount=1)
    i += 1
wb.save("new_data.xlsx")

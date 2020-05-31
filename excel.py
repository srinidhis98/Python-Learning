import openpyxl as xl


def compare_date(workbook, sheet_name1, sheet_name2):
    cell_list = []
    sheet1 = workbook[sheet_name1]
    sheet2 = workbook[sheet_name2]
    print(sheet_name1 + " " + sheet_name2)
    print(sheet1.max_row)
    print(sheet2.max_row)
    row_num = 2
    row = 2
    while row_num <= 10:
        cell_element = sheet2.cell(row_num, 2)
        cell_list.insert(0, cell_element.value)
        cell_element = sheet2.cell(row_num, 3)
        cell_list.insert(1, cell_element.value)
        print(cell_list)
        cell1 = sheet1.cell(row, 1)
        print(type(cell1.value))
        cell2 = sheet2.cell(row, 1)
        print(type(cell2.value))
        if cell1.value == cell2.value:
            # print('vmstat time == handoff time')
            append_cell = sheet1.cell(row, 6)
            append_cell.value = cell_list[0]
            append_cell = sheet1.cell(row, 7)
            append_cell.value = cell_list[1]
            row += 1

        # while cell1.value != cell2.value:
        while cell1.value < cell2.value:
            row += 1
            cell1 = sheet1.cell(row, 1)
            if cell1.value == cell2.value:
                # print('vmstat time == handoff time')
                append_cell = sheet1.cell(row, 6)
                append_cell.value = cell_list[0]
                append_cell = sheet1.cell(row, 7)
                append_cell.value = cell_list[1]
                row += 1
            elif cell1.value != cell2.value:
                row += 1
                sheet1.insert_rows(row, amount=1)
                append_cell = sheet1.cell(row, 6)
                append_cell.value = cell_list[0]
                append_cell = sheet1.cell(row, 7)
                append_cell.value = cell_list[1]
            else:
                pass

        while cell1.value > cell2.value:
            row -= 1
            cell1 = sheet1.cell(row, 1)
            if cell1.value == cell2.value:
                # print('vmstat time == handoff time')
                append_cell = sheet1.cell(row, 6)
                append_cell.value = cell_list[0]
                append_cell = sheet1.cell(row, 7)
                append_cell.value = cell_list[1]
                row += 1
            elif cell1.value < cell2.value:
                row -= 1
                sheet1.insert_rows(row, amount=1)
                append_cell = sheet1.cell(row, 6)
                append_cell.value = cell_list[0]
                append_cell = sheet1.cell(row, 7)
                append_cell.value = cell_list[1]
            else:
                pass

        row_num += 1
        cell_list.clear()


def main():
    wb = xl.load_workbook("dataa2.xlsx")
    # status = delete_rows(wb, "handoff")
    # print(status)
    compare_date(wb, "flow handoff", "mpstat")
    wb.save("new_dataa2.xlsx")


if __name__ == '__main__':
    main()
